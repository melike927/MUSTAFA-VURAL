import re
import time
import random
import unicodedata
from urllib.parse import quote_plus, unquote

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

EXCEL_PATH = r"c:\Users\melike\Desktop\TÜBİTAK 3501 için.xlsx"
BASE_AVESIS = "https://avesis.omu.edu.tr"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"

TR_MAP = str.maketrans({
    "ç": "c", "Ç": "c", "ğ": "g", "Ğ": "g", "ı": "i", "İ": "i",
    "ö": "o", "Ö": "o", "ş": "s", "Ş": "s", "ü": "u", "Ü": "u",
})


def normalize_text(value: str) -> str:
    t = value.translate(TR_MAP)
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"[^a-zA-Z\s]", " ", t).lower()
    t = re.sub(r"\s+", " ", t).strip()
    return t


def score_name_match(target_name: str, page_text: str) -> int:
    target_tokens = set(normalize_text(target_name).split())
    page_tokens = set(normalize_text(page_text).split())
    if not target_tokens:
        return 0
    return len(target_tokens & page_tokens)


def get_search_result_links(session: requests.Session, person_name: str):
    # DuckDuckGo HTML endpoint is usually less strict for scripted reads.
    query = f'site:avesis.omu.edu.tr "{person_name}"'
    url = f"https://duckduckgo.com/html/?q={quote_plus(query)}&kl=tr-tr"
    try:
        r = session.get(url, timeout=30)
    except requests.RequestException:
        return []
    if r.status_code != 200:
        return []

    soup = BeautifulSoup(r.text, "lxml")
    links = []

    # DuckDuckGo result links are commonly in a.result__a and may include redirect params.
    for a in soup.select("a[href]"):
        href = a.get("href", "")
        if "uddg=" in href:
            target = href.split("uddg=", 1)[1].split("&", 1)[0]
            target = unquote(target)
            if target.startswith(BASE_AVESIS):
                links.append(target)
        elif href.startswith(BASE_AVESIS):
            links.append(href)

    # Keep likely profile pages and unique order.
    seen = set()
    filtered = []
    blocked_parts = [
        "/yayin", "/arastirma-grubu", "/iletisim", "/hakkinda", "/arama",
        "/arastirmaci-arama", "/detayli-arama", "/researcher", "/journal",
        "/unitreport", "/surdurulebilirlik", "/yapay-zeka"
    ]
    for link in links:
        l = link.strip().rstrip("/")
        if not l.startswith(BASE_AVESIS):
            continue
        path = l.replace(BASE_AVESIS, "")
        if not path or path == "":
            continue
        if any(part in path for part in blocked_parts):
            continue
        # Prefer short paths like /ad.soyad
        if path.count("/") > 1:
            continue
        if l not in seen:
            seen.add(l)
            filtered.append(l)

    return filtered


def extract_email_from_profile(session: requests.Session, profile_url: str):
    try:
        r = session.get(profile_url, timeout=30)
    except requests.RequestException:
        return None, None

    if r.status_code != 200:
        return None, None

    text = r.text
    if "Sayfa Bulunamadı" in text:
        return None, None

    # First, strict OMU domain match.
    emails = re.findall(r"[A-Za-z0-9._%+-]+@omu\.edu\.tr", text, flags=re.IGNORECASE)
    if emails:
        return sorted(set(e.lower() for e in emails))[0], text

    # Fallback: mailto link.
    soup = BeautifulSoup(text, "lxml")
    for a in soup.select("a[href^='mailto:']"):
        mail = a.get("href", "").replace("mailto:", "").strip().lower()
        if mail.endswith("@omu.edu.tr"):
            return mail, text

    return None, text


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        name_col = headers.index("Adı") + 1
    except ValueError:
        raise RuntimeError("Excel dosyasında 'Adı' sütunu bulunamadı.")

    # Reuse first empty header columns as output columns.
    mail_col = None
    source_col = None
    status_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value is None:
            if mail_col is None:
                mail_col = c
            elif source_col is None:
                source_col = c
            elif status_col is None:
                status_col = c
                break

    if mail_col is None:
        mail_col = ws.max_column + 1
    if source_col is None:
        source_col = mail_col + 1
    if status_col is None:
        status_col = source_col + 1

    ws.cell(1, mail_col).value = "Mail"
    ws.cell(1, source_col).value = "Kaynak URL"
    ws.cell(1, status_col).value = "Durum"

    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "tr-TR,tr;q=0.9,en;q=0.8"})

    total = ws.max_row - 1
    found_count = 0

    for r in range(2, ws.max_row + 1):
        person_name = ws.cell(r, name_col).value
        if not person_name:
            ws.cell(r, status_col).value = "isim-yok"
            continue

        # If already filled, skip.
        existing = ws.cell(r, mail_col).value
        if existing:
            ws.cell(r, status_col).value = "zaten-var"
            continue

        links = get_search_result_links(session, str(person_name))

        best_email = None
        best_url = None
        best_score = 0

        for link in links[:5]:
            email, page_text = extract_email_from_profile(session, link)
            if page_text is None:
                continue
            score = score_name_match(str(person_name), page_text)
            if email and score >= best_score:
                best_email = email
                best_url = link
                best_score = score

            time.sleep(random.uniform(0.4, 0.9))

        if best_email:
            ws.cell(r, mail_col).value = best_email
            ws.cell(r, source_col).value = best_url
            ws.cell(r, status_col).value = "bulundu"
            found_count += 1
        else:
            ws.cell(r, status_col).value = "bulunamadi"

        if (r - 1) % 20 == 0:
            print(f"ilerleme: {r-1}/{total} satir, bulunan={found_count}")
            wb.save(EXCEL_PATH)

        time.sleep(random.uniform(0.7, 1.4))

    wb.save(EXCEL_PATH)
    print(f"tamamlandi: toplam={total}, bulunan={found_count}")


if __name__ == "__main__":
    main()
