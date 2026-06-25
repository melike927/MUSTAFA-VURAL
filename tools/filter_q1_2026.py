from copy import copy
from datetime import datetime
from openpyxl import load_workbook, Workbook

source = r"c:\Users\melike\AppData\Local\Packages\5319275A.WhatsAppDesktop_cv1g1gvanyjgm\LocalState\sessions\A2C42D50E1A26C42297E87442331A5C168642750\transfers\2026-14\İş Destek Verilenler.xlsx"
outfile = r"c:\Users\melike\AppData\Local\Packages\5319275A.WhatsAppDesktop_cv1g1gvanyjgm\LocalState\sessions\A2C42D50E1A26C42297E87442331A5C168642750\transfers\2026-14\İş Destek Verilenler_2026_ilk_uc_ay.xlsx"

wb = load_workbook(source)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
date_col = headers.index("Destek Tarihi") + 1

filtered_rows = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, date_col).value
    if not hasattr(v, "year") or not hasattr(v, "month"):
        continue
    if v.year == 2026 and v.month in (1, 2, 3):
        filtered_rows.append(r)

new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Modül 2"

# Copy column widths
for col_letter, dim in ws.column_dimensions.items():
    new_ws.column_dimensions[col_letter].width = dim.width
    new_ws.column_dimensions[col_letter].hidden = dim.hidden

# Copy header row with style
for c in range(1, ws.max_column + 1):
    src = ws.cell(1, c)
    dst = new_ws.cell(1, c, src.value)
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)

# Copy matching rows
out_row = 2
for r in filtered_rows:
    for c in range(1, ws.max_column + 1):
        src = ws.cell(r, c)
        dst = new_ws.cell(out_row, c, src.value)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    out_row += 1

# Freeze and filter
new_ws.freeze_panes = "A2"
new_ws.auto_filter.ref = f"A1:{chr(64 + ws.max_column)}{new_ws.max_row}"

new_wb.save(outfile)
print("saved", outfile)
print("rows", len(filtered_rows))
