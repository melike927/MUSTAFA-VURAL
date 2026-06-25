import re
import unicodedata
import requests
from openpyxl import load_workbook

EXCEL = r"c:\Users\melike\Desktop\TÜBİTAK 3501 için.xlsx"

TR_MAP = str.maketrans({
    "ç": "c", "Ç": "c", "ğ": "g", "Ğ": "g", "ı": "i", "İ": "i",
    "ö": "o", "Ö": "o", "ş": "s", "Ş": "s", "ü": "u", "Ü": "u",
})


def normalize_name(name: str) -> str:
    t = name.translate(TR_MAP)
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"[^a-zA-Z\s]", " ", t).lower()
    t = re.sub(r"\s+", " ", t).strip()
    return t


def candidate_slugs(full_name: str):
    parts = normalize_name(full_name).split()
    if len(parts) < 2:
        return []
    first = parts[0]
    last = parts[-1]
    middle = parts[1:-1]
    c = []
    c.append(f"{first}.{last}")
    c.append(f"{first}{last}")
    c.append(f"{first[0]}{last}")
    if middle:
        c.append(".".join([first] + middle + [last]))
        c.append("".join([first] + middle + [last]))
        c.append(f"{first}.{middle[0]}.{last}")
    # unique preserve order
    seen = set()
    out = []
    for x in c:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out

wb = load_workbook(EXCEL)
ws = wb.active
names = [ws.cell(r, 2).value for r in range(2, min(ws.max_row, 60)+1)]
session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0"})

for name in names[:20]:
    found = None
    for slug in candidate_slugs(name):
        url = f"https://avesis.omu.edu.tr/{slug}"
        try:
            r = session.get(url, timeout=15, allow_redirects=True)
        except Exception:
            continue
        if r.status_code == 200 and "Sayfa Bulunamadı" not in r.text and "AVESİS" in r.text:
            m = re.search(r"[A-Za-z0-9._%+-]+@omu\.edu\.tr", r.text)
            found = (slug, m.group(0) if m else None)
            break
    print(name, "=>", found)
