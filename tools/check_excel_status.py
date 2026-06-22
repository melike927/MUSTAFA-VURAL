from collections import Counter
from openpyxl import load_workbook

p = r"c:\Users\melike\Desktop\TÜBİTAK 3501 için.xlsx"
ws = load_workbook(p).active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print("headers:", headers)

mail_col = headers.index("Mail") + 1 if "Mail" in headers else None
durum_col = headers.index("Durum") + 1 if "Durum" in headers else None
print("mail_col", mail_col, "durum_col", durum_col)

filled = 0
status_counter = Counter()
if mail_col:
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, mail_col).value:
            filled += 1
        if durum_col:
            status_counter[str(ws.cell(r, durum_col).value)] += 1

print("rows", ws.max_row - 1, "mail_filled", filled)
print("status_counts", dict(status_counter))
