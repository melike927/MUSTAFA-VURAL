from openpyxl import load_workbook
from collections import Counter
from datetime import datetime

p = r"c:\Users\melike\AppData\Local\Packages\5319275A.WhatsAppDesktop_cv1g1gvanyjgm\LocalState\sessions\A2C42D50E1A26C42297E87442331A5C168642750\transfers\2026-14\İş Destek Verilenler.xlsx"
ws = load_workbook(p).active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
col = headers.index("Destek Tarihi") + 1
print("date_col", col)
vals = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, col).value
    vals.append(v)
    if r <= 10:
        print(r, repr(v), type(v))

counts = Counter()
for v in vals:
    if hasattr(v, "year") and hasattr(v, "month"):
        counts[(v.year, v.month)] += 1
    else:
        s = str(v)
        for fmt in ["%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y %H:%M:%S"]:
            try:
                dt = datetime.strptime(s, fmt)
                counts[(dt.year, dt.month)] += 1
                break
            except Exception:
                pass
        else:
            counts[("unparsed", s)] += 1
print("counts", counts.most_common(20))
